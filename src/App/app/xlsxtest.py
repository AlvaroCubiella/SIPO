from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from PyPDF2 import PdfFileWriter, PdfFileReader
from utils import xlsxStructura
from tempfile import NamedTemporaryFile 
import os

def create_excel_with_merged_cells():
    # Crear un nuevo libro de trabajo
    wb = Workbook()
    # Seleccionar la hoja activa
    ws = wb.active

     # Configurar la orientación horizontal (horizontal = True) o vertical (horizontal = False)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # Orientación horizontal

    # Configurar los márgenes de la hoja en milímetros
    ws.page_margins.left = 0.2          # Margen izquierdo en milímetros
    ws.page_margins.right = 0.2         # Margen derecho en milímetros
    ws.page_margins.top = 0.5           # Margen superior en milímetros
    ws.page_margins.bottom = 0.5        # Margen inferior en milímetros

    range_ascii = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L',
                   'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X',
                   'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 
                   'AI', 'AJ']
    
    for i in range_ascii:
        ws.column_dimensions[i].width = 3.67

    ws.row_dimensions[1].height = 37
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 5
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 25
    ws.row_dimensions[7].height = 5

    celdas = xlsxStructura()


    border = Border(left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')) 
    
    # Definir la alineación para ajustar el texto
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    # Definir el estilo de la fuente
    font = Font(name='Arial', size=12, bold=True)

    # Combinar celdas para la cabecera de la campaña
    for key, value in celdas.items():
        # Obtener el rango de celdas
        ini_col = value['Range']['Col']['Ini']
        fin_col = value['Range']['Col']['Fin']
        ini_fila = value['Range']['Fila']
        lbl = value['lbl']
        ini = f"{ini_col}{ini_fila}"
        fin = f"{fin_col}{ini_fila}"
        range_cells = ws[ini:fin]
        # Combino celda del nombre
        ws.merge_cells(f"{ini_col}{ini_fila}:{fin_col}{ini_fila}")
        ws[ini] = value['lbl']
        
        # Aplicar el borde al rango de celdas combinadas
        for row in ws[f"{ini_col}{ini_fila}:{fin_col}{ini_fila}"]:
            for cell in row:
                cell.border = border
                cell.font = font
                cell.alignment = alignment
        if 'campo' in celdas[key]:
            if celdas[key]['campo']==True:
                # Genero un campo para completar
                ws.merge_cells(f"{ini_col}{ini_fila+1}:{fin_col}{ini_fila+1}")
                for row in ws[f"{ini_col}{ini_fila+1}:{fin_col}{ini_fila+1}"]:
                    for cell in row:
                        cell.border = border
        else:
            print(key)


    pdf_file = "hoja_de_calculo_con_combinacion.pdf"
    temp_dir = os.getcwd()


    # Guardar el archivo Excel temporalmente como PDF
    with NamedTemporaryFile(suffix='.pdf', dir=temp_dir, delete=False) as tmp_excel_pdf:
        wb.save(tmp_excel_pdf.name)

    wb.save(tmp_excel_pdf.name)

     # Copiar el contenido del archivo PDF temporal a un nuevo archivo PDF
    pdf_writer = PdfFileWriter()
    pdf_reader = PdfFileReader(tmp_excel_pdf.name)
    for page in range(pdf_reader.numPages):
        pdf_writer.addPage(pdf_reader.getPage(page))

    # Guardar el archivo PDF final
    with open(pdf_file, 'wb') as f:
        pdf_writer.write(f)

    # Guardar el archivo Excel
    wb.save("hoja_de_calculo_con_combinacion.xlsx")

if __name__ == "__main__":
    create_excel_with_merged_cells()