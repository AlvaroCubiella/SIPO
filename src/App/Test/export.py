import pandas as pd
import json
from openpyxl import load_workbook


# Cargar el archivo JSON
file = "E:\\SiavoVB_Anexo\\Estructura\\VA\\2024\\003\\Varios\\VA202403.json"
with open(file, 'r') as f:
    data = json.load(f)

# Crear un DataFrame con los datos de las estaciones
estaciones_data = []
for estacion, detalles in data['Estaciones'].items():
    estaciones_data.append({
        'Estacion': estacion,
        'Latitud Inicial': detalles['Posicion']['Inicio']['Latitud'],
        'Longitud Inicial': detalles['Posicion']['Inicio']['Longitud'],
        'Latitud Final': detalles['Posicion']['Fin']['Latitud'],
        'Longitud Final': detalles['Posicion']['Fin']['Longitud']
        # Agrega más campos según sea necesario
    })

# Crear un DataFrame con los datos de las estaciones
df_estaciones = pd.DataFrame(estaciones_data)

# Cargar la plantilla de Excel
plantilla_path = 'E:\\Programas\\Py\\Planilla_Estaciones\\Planilla_V1.xlsx'
wb = load_workbook(plantilla_path)

# Obtener la hoja de cálculo donde quieres escribir los datos
ws = wb['Hoja1']

# Escribir los datos en la hoja de cálculo
for index, row in df_estaciones.iterrows():
    for col, value in row.items():
        ws[f'{col}{index + 2}'] = value

# Guardar el archivo Excel
wb.save('estaciones_completas.xlsx')